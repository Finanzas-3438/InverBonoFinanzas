import openpyxl
from io import BytesIO
from django.http import HttpResponse
import pandas as pd
from django.shortcuts import get_object_or_404
from .models import Bond

def generate_bond_excel(bond_id):
    bond = get_object_or_404(Bond, pk=bond_id)
    buffer = BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Detalle del Bono'
    col_widths = [6, 24, 18, 2, 6, 24, 18]
    for col, width in enumerate(col_widths, 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        worksheet.column_dimensions[col_letter].width = width

    # Main title styles
    title_style = {
        'font': {'bold': True, 'size': 18, 'color': 'FFFFFF'},
        'fill': {'fgColor': '2C3E50', 'patternType': 'solid'},
        'alignment': {'horizontal': 'center', 'vertical': 'center'}
    }
    
    # Section header styles
    data_header_style = {
        'font': {'bold': True, 'size': 14, 'color': 'FFFFFF'},
        'fill': {'fgColor': '16A085', 'patternType': 'solid'},
        'alignment': {'horizontal': 'center', 'vertical': 'center', 'textRotation': 90}
    }
    
    results_header_style = {
        'font': {'bold': True, 'size': 14, 'color': 'FFFFFF'},
        'fill': {'fgColor': '2980B9', 'patternType': 'solid'},
        'alignment': {'horizontal': 'center', 'vertical': 'center', 'textRotation': 90}
    }
    
    # Subsection styles
    data_subheader_style = {
        'font': {'bold': True, 'size': 12, 'color': 'FFFFFF'},
        'fill': {'fgColor': '27AE60', 'patternType': 'solid'},
        'border': {'left': {'style': 'thin', 'color': '000000'},
                  'right': {'style': 'thin', 'color': '000000'}, 
                  'top': {'style': 'thin', 'color': '000000'}, 
                  'bottom': {'style': 'thin', 'color': '000000'}},
        'alignment': {'horizontal': 'center', 'vertical': 'center'}
    }
    
    results_subheader_style = {
        'font': {'bold': True, 'size': 12, 'color': 'FFFFFF'},
        'fill': {'fgColor': '3498DB', 'patternType': 'solid'},
        'border': {'left': {'style': 'thin', 'color': '000000'},
                  'right': {'style': 'thin', 'color': '000000'}, 
                  'top': {'style': 'thin', 'color': '000000'}, 
                  'bottom': {'style': 'thin', 'color': '000000'}},
        'alignment': {'horizontal': 'center', 'vertical': 'center'}
    }
    
    # Data section styles
    data_label_style = {
        'font': {'bold': True, 'size': 10, 'color': '000000'},
        'fill': {'fgColor': 'A9DFBF', 'patternType': 'solid'},
        'border': {'left': {'style': 'thin', 'color': '000000'},
                  'right': {'style': 'thin', 'color': '000000'}, 
                  'top': {'style': 'thin', 'color': '000000'}, 
                  'bottom': {'style': 'thin', 'color': '000000'}},
        'alignment': {'horizontal': 'left', 'vertical': 'center'}
    }
    
    data_value_style = {
        'font': {'bold': False, 'size': 10, 'color': '000000'},
        'fill': {'fgColor': 'EAFAF1', 'patternType': 'solid'},
        'border': {'left': {'style': 'thin', 'color': '000000'},
                  'right': {'style': 'thin', 'color': '000000'},
                  'top': {'style': 'thin', 'color': '000000'}, 
                  'bottom': {'style': 'thin', 'color': '000000'}},
        'alignment': {'horizontal': 'left', 'vertical': 'center'}
    }
    
    # Results section styles
    results_label_style = {
        'font': {'bold': True, 'size': 10, 'color': '000000'},
        'fill': {'fgColor': 'AED6F1', 'patternType': 'solid'},
        'border': {'left': {'style': 'thin', 'color': '000000'},
                  'right': {'style': 'thin', 'color': '000000'}, 
                  'top': {'style': 'thin', 'color': '000000'}, 
                  'bottom': {'style': 'thin', 'color': '000000'}},
        'alignment': {'horizontal': 'left', 'vertical': 'center'}
    }
    
    results_value_style = {
        'font': {'bold': False, 'size': 10, 'color': '000000'},
        'fill': {'fgColor': 'EBF5FB', 'patternType': 'solid'},
        'border': {'left': {'style': 'thin', 'color': '000000'},
                  'right': {'style': 'thin', 'color': '000000'}, 
                  'top': {'style': 'thin', 'color': '000000'}, 
                  'bottom': {'style': 'thin', 'color': '000000'}},
        'alignment': {'horizontal': 'left', 'vertical': 'center'}
    }
    
    footer_style = {
        'font': {'italic': True, 'size': 9, 'color': '000000'},
        'fill': {'fgColor': 'EAEDED', 'patternType': 'solid'},
        'border': {'left': {'style': 'thin', 'color': '000000'},
                  'right': {'style': 'thin', 'color': '000000'}, 
                  'top': {'style': 'thin', 'color': '000000'}, 
                  'bottom': {'style': 'thin', 'color': '000000'}},
        'alignment': {'horizontal': 'right', 'vertical': 'center'}
    }
    
    # Estilo para subtítulo
    subtitle_style = {
        'font': {'bold': True, 'size': 14, 'color': 'FFFFFF'},
        'fill': {'fgColor': '16A085', 'patternType': 'solid'},
        'border': {'left': {'style': 'medium', 'color': '000000'},
                  'right': {'style': 'medium', 'color': '000000'}, 
                  'top': {'style': 'medium', 'color': '000000'}, 
                  'bottom': {'style': 'medium', 'color': '000000'}},
        'alignment': {'horizontal': 'center', 'vertical': 'center'}
    }
    
    # Apply style function with fixed border application
    def apply_style(cell, style_dict):
        if 'font' in style_dict:
            font_params = {
                'bold': style_dict['font'].get('bold', False),
                'size': style_dict['font'].get('size', 11),
                'color': style_dict['font'].get('color', '000000'),
                'italic': style_dict['font'].get('italic', False)
            }
            cell.font = openpyxl.styles.Font(**font_params)
            
        if 'fill' in style_dict:
            cell.fill = openpyxl.styles.PatternFill(
                start_color=style_dict['fill'].get('fgColor', 'FFFFFF'),
                fill_type=style_dict['fill'].get('patternType', 'solid')
            )
        
        if 'border' in style_dict:
            border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(
                    style=style_dict['border']['left']['style'],
                    color=style_dict['border']['left']['color']
                ),
                right=openpyxl.styles.Side(
                    style=style_dict['border']['right']['style'],
                    color=style_dict['border']['right']['color']
                ),
                top=openpyxl.styles.Side(
                    style=style_dict['border']['top']['style'],
                    color=style_dict['border']['top']['color']
                ),
                bottom=openpyxl.styles.Side(
                    style=style_dict['border']['bottom']['style'],
                    color=style_dict['border']['bottom']['color']
                )
            )
            cell.border = border
            
        if 'alignment' in style_dict:
            alignment_params = {
                'horizontal': style_dict['alignment'].get('horizontal', 'general'),
                'vertical': style_dict['alignment'].get('vertical', 'bottom'),
                'wrap_text': style_dict['alignment'].get('wrap_text', False)
            }
            if 'textRotation' in style_dict['alignment']:
                alignment_params['text_rotation'] = style_dict['alignment']['textRotation']
            
            cell.alignment = openpyxl.styles.Alignment(**alignment_params)
    
    # Title and subtitle
    worksheet.merge_cells('A1:G1')
    title_cell = worksheet['A1']
    title_cell.value = f"DETALLE DEL BONO: {bond.name}"
    apply_style(title_cell, title_style)
    worksheet.row_dimensions[1].height = 30
    
    worksheet.merge_cells('A2:G2')
    subtitle_cell = worksheet['A2']
    subtitle_cell.value = f"Emitido el {bond.issue_date.strftime('%d/%m/%Y')}"
    apply_style(subtitle_cell, subtitle_style)
    worksheet.row_dimensions[2].height = 22
    
    start_data_row = 5
    
    data_row = start_data_row
    results_row = start_data_row
    
    # --- COLUMNA DE DATOS ---
    
    # Sección 1: Del Bono
    worksheet.merge_cells(f'B{data_row}:C{data_row}')
    data_subheader = worksheet[f'B{data_row}']
    data_subheader.value = "Del Bono"
    apply_style(data_subheader, data_subheader_style)
    worksheet.row_dimensions[data_row].height = 20
    data_row += 1
    
    # Datos del Bono
    bond_data = [
        ["Valor Nominal", f"S/ {bond.nominal_value:.2f}"],
        ["Valor Comercial", f"S/ {bond.commercial_value:.2f}"],
        ["Nº de Años", f"{bond.years_number}"],
        ["Frecuencia del cupón", bond.get_coupon_frequency_display()],
        ["Días x Año", "360"],
        ["Tipo de Tasa de Interés", bond.get_interest_rate_type_display()],
        ["Capitalización", bond.get_capitalization_display()],
        ["Tasa de interés", f"{bond.interest_rate}%"],
        ["Tasa anual de descuento", f"{bond.annual_discount_rate}%"],
        ["Imp. a la Renta", f"{bond.income_tax}%"],
        ["Fecha de Emisión", bond.issue_date.strftime('%d/%m/%Y')]
    ]
    
    for item in bond_data:
        label_cell = worksheet.cell(row=data_row, column=2, value=item[0])
        apply_style(label_cell, data_label_style)
        value_cell = worksheet.cell(row=data_row, column=3, value=item[1])
        apply_style(value_cell, data_value_style)
        worksheet.row_dimensions[data_row].height = 18
        data_row += 1
    
    # Sección 2: De los Gastos Iniciales
    worksheet.merge_cells(f'B{data_row}:C{data_row}')
    expenses_header = worksheet[f'B{data_row}']
    expenses_header.value = "De los Gastos Iniciales"
    apply_style(expenses_header, data_subheader_style)
    worksheet.row_dimensions[data_row].height = 20
    data_row += 1
    
    # Datos de gastos iniciales
    expenses_data = [
        ["% Prima", f"{bond.premium_percentage}%"],
        ["% Estructuración", f"{bond.structuring_percentage}% ({bond.get_structuring_type_display()})"],
        ["% Colocación", f"{bond.placement_percentage}% ({bond.get_placement_type_display()})"],
        ["% Flotación", f"{bond.float_percentage}% ({bond.get_float_type_display()})"],
        ["% CAVALI", f"{bond.cavali_percentage}% ({bond.get_cavali_type_display()})"]
    ]
    
    for item in expenses_data:
        label_cell = worksheet.cell(row=data_row, column=2, value=item[0])
        apply_style(label_cell, data_label_style)
        value_cell = worksheet.cell(row=data_row, column=3, value=item[1])
        apply_style(value_cell, data_value_style)
        worksheet.row_dimensions[data_row].height = 18
        data_row += 1
    
    # --- COLUMNA DE RESULTADOS ---
    
    # Sección 1: De la Estructuración del Bono
    worksheet.merge_cells(f'F{results_row}:G{results_row}')
    results_subheader = worksheet[f'F{results_row}']
    results_subheader.value = "De la Estructuración del Bono"
    apply_style(results_subheader, results_subheader_style)
    worksheet.row_dimensions[results_row].height = 20
    results_row += 1
    
    # Resultados de la estructuración
    structure_results = [
        ["Días capitalización", str(bond.capitalization_days)],
        ["Nº Períodos por Año", str(bond.periods_per_year)],
        ["Nº Total de Periodos", str(bond.total_periods)],
        ["Tasa efectiva anual", f"{bond.effective_annual_rate*100:.3f}%"],
        ["Tasa efectiva {0}".format(bond.get_coupon_frequency_display()), f"{bond.effective_coupon_rate*100:.3f}%"],
        ["COK {0}".format(bond.get_coupon_frequency_display()), f"{bond.cok*100:.3f}%"],
        ["Costes Iniciales Emisor", f"S/ {bond.issuer_initial_cost:.2f}"],
        ["Costes Iniciales Bonista", f"S/ {bond.bondholder_initial_cost:.2f}"]
    ]
    
    for item in structure_results:
        result_label_cell = worksheet.cell(row=results_row, column=6, value=item[0])
        apply_style(result_label_cell, results_label_style)
        result_value_cell = worksheet.cell(row=results_row, column=7, value=item[1])
        apply_style(result_value_cell, results_value_style)
        worksheet.row_dimensions[results_row].height = 18
        results_row += 1
    
    # Sección 2: Del Precio Actual y Utilidad
    worksheet.merge_cells(f'F{results_row}:G{results_row}')
    price_header = worksheet[f'F{results_row}']
    price_header.value = "Del Precio Actual y Utilidad"
    apply_style(price_header, results_subheader_style)
    worksheet.row_dimensions[results_row].height = 20
    results_row += 1
    
    # Datos de precio y utilidad
    price_results = [
        ["Precio Actual", f"S/ {bond.precio_actual:.2f}"],
        ["Utilidad / Pérdida", f"S/ {bond.utilidad:.2f}"]
    ]
    
    for item in price_results:
        result_label_cell = worksheet.cell(row=results_row, column=6, value=item[0])
        apply_style(result_label_cell, results_label_style)
        result_value_cell = worksheet.cell(row=results_row, column=7, value=item[1])
        apply_style(result_value_cell, results_value_style)
        worksheet.row_dimensions[results_row].height = 18
        results_row += 1
    
    # Sección 3: De los Ratios de Decisión
    worksheet.merge_cells(f'F{results_row}:G{results_row}')
    risk_header = worksheet[f'F{results_row}']
    risk_header.value = "De los Ratios de Decisión"
    apply_style(risk_header, results_subheader_style)
    worksheet.row_dimensions[results_row].height = 20
    results_row += 1
    
    # Datos de ratios de decisión
    risk_data = [
        ["Duración", f"{bond.duracion:.2f}"],
        ["Convexidad", f"{bond.convexidad:.2f}"],
        ["Total", f"{bond.total:.2f}"],
        ["Duración Modificada", f"{bond.duracion_modificada:.2f}"]
    ]
    
    for item in risk_data:
        result_label_cell = worksheet.cell(row=results_row, column=6, value=item[0])
        apply_style(result_label_cell, results_label_style)
        result_value_cell = worksheet.cell(row=results_row, column=7, value=item[1])
        apply_style(result_value_cell, results_value_style)
        worksheet.row_dimensions[results_row].height = 18
        results_row += 1
    
    # Sección 4: De los Indicadores de Rentabilidad
    worksheet.merge_cells(f'F{results_row}:G{results_row}')
    profitability_header = worksheet[f'F{results_row}']
    profitability_header.value = "De los Indicadores de Rentabilidad"
    apply_style(profitability_header, results_subheader_style)
    worksheet.row_dimensions[results_row].height = 20
    results_row += 1
    
    # Datos de rentabilidad
    profitability_data = [
        ["TCEA Emisor", f"{bond.tcea_emisor*100:.3f}%"],
        ["TCEA Emisor c/Escudo", f"{bond.tcea_emisor_escudo*100:.3f}%"],
        ["TREA Bonista", f"{bond.trea_bonista*100:.3f}%"]
    ]
    
    for item in profitability_data:
        result_label_cell = worksheet.cell(row=results_row, column=6, value=item[0])
        apply_style(result_label_cell, results_label_style)
        result_value_cell = worksheet.cell(row=results_row, column=7, value=item[1])
        apply_style(result_value_cell, results_value_style)
        worksheet.row_dimensions[results_row].height = 18
        results_row += 1
    

    data_end_row = data_row - 1
    results_end_row = results_row - 1

    # Merge cells for main headers
    worksheet.merge_cells(f'A{start_data_row}:A{data_end_row}')
    data_main_header = worksheet[f'A{start_data_row}']
    data_main_header.value = "D A T O S"
    apply_style(data_main_header, data_header_style)
    
    worksheet.merge_cells(f'E{start_data_row}:E{results_end_row}')
    results_main_header = worksheet[f'E{start_data_row}']
    results_main_header.value = "R E S U L T A D O S"
    apply_style(results_main_header, results_header_style)
    
    # Pie de página
    current_row = max(data_end_row, results_end_row) + 1
    worksheet.merge_cells(f'A{current_row}:G{current_row}')
    note_cell = worksheet[f'A{current_row}']
    note_cell.value = f"InverBono - Exportado el {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}"
    apply_style(note_cell, footer_style)
    
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=bono_{bond.name.replace(' ', '_')}.xlsx'
    return response
